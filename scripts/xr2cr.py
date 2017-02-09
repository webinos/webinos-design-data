#!/usr/bin/python

import os
from openpyxl import load_workbook


if __name__ == "__main__":
 
  outputDir = os.environ['TMP_DIR']
  reqDir = os.environ['REQ_DIR']
  priorityLookup = {}
  priorityLookup['Low'] = '3'
  priorityLookup['Medium'] = '2'
  priorityLookup['High'] = '1'

  cellDict = {}
  cellDict[1] = 'Name'
  cellDict[2] = 'Originator'
  cellDict[3] = 'Functional area'
  cellDict[4] = 'Priority'
  cellDict[5] = 'Description'
  cellDict[6] = 'Fit Criterion'
  cellDict[7] = 'Comments'

  xmlBuf = '<?xml version="1.0"?>\n<!DOCTYPE goals PUBLIC "-//CAIRIS//DTD GOALS USABILITY 1.0//EN" "http://cairis.org/dtd/goals.dtd">\n\n<goals>\n\n'
  for fileName in ['TMS.xlsx','PS.xlsx','NM.xlsx','NC.xlsx','LC.xlsx','ID.xlsx','DA.xlsx','CAP.xlsx']:
    wb = load_workbook(reqDir + '/' + fileName)
    ws = wb.get_sheet_by_name('In scope')
    reqLabel = 1
    for row in ws.iter_rows(min_row = 2):
      rowDict = {}
      for cell in row:
        if cell.col_idx < 8:
          reqAttribute = cellDict[cell.col_idx]
          rowDict[reqAttribute] = cell.value
      xmlBuf += '<requirement name=\"' + rowDict['Name'] + '\" reference=\"' + rowDict['Functional area'] + '\" reference_type=\"environment\" label=\"' + str(reqLabel) + '\" type=\"Functional\" priority=\"' + priorityLookup[rowDict['Priority']] + '\">\n  <description>' + rowDict['Description'] + '</description>\n  <rationale>' + rowDict['Comments'] + '</rationale>\n  <fit_criterion>' + rowDict['Fit Criterion'] + '</fit_criterion>\n  <originator>' + rowDict['Originator'] + '</originator>\n</requirement>\n'
      reqLabel += 1
  xmlBuf += '</goals>'

  outputFile = outputDir + '/requirements.xml'
  f = open(outputFile,'w')
  f.write(xmlBuf)
  f.close() 
