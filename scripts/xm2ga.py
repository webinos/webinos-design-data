#!/usr/bin/python

# Converts XLS  obstacle mitigation sheets to a cairis model of domain properties and goal associations

import os
from openpyxl import load_workbook

if __name__ == "__main__":
 
  outputDir = os.environ['TMP_DIR']
  archDir = os.environ['ARCHITECTURE_DIR']


  cellDict = {}
  cellDict[1] = 'Obstacle'
  cellDict[2] = 'Goal'
  cellDict[3] = 'Definition'
  cellDict[4] = 'Affected Components'
  cellDict[5] = 'Satisfied'
  cellDict[6] = 'Rationale'

  xmlBuf = '<?xml version="1.0"?>\n<!DOCTYPE cairis_model PUBLIC "-//CAIRIS//DTD MODEL 1.0//EN" "http://cairis.org/dtd/cairis_model.dtd">\n\n<cairis_model>\n\n'
  dpBuf = ''
  gBuf = ''
  gaBuf = ''
  aaBuf = '|_.Obstacle|_.Mitigating Requirement Name|_.Mitigating Requirement Definition|_.Affected Components|_.Satisfied (Y/N)|_.Rationale|\n'

  apMitFile = 'attackPatternMitigations.xlsx'
  wb = load_workbook(archDir + '/' + apMitFile)
  ws = wb.get_sheet_by_name('Goals')

  for row in ws.iter_rows(min_row = 2):
    rowDict = {}
    for cell in row:
      if cell.col_idx < 7:
        reqAttribute = cellDict[cell.col_idx]
        rowDict[reqAttribute] = cell.value

    obsNames = rowDict['Obstacle']
    gdpName = rowDict['Goal']
    gdpDef = rowDict['Definition']
    cNames = rowDict['Affected Components']
    isSatisfied = rowDict['Satisfied']
    dpDef = rowDict['Rationale']

    if isSatisfied == 'N':
      dpBuf += '<domainproperty name=\"' + gdpName + '\" type="Invariant" originator="WP 3 workshop - August 2012">\n  <definition>' + dpDef + '</definition>\n</domainproperty>\n'
    else:
      gBuf += '<goal name=\"' + gdpName + '\" originator="WP 3 workshop - August 2012">\n  <goal_environment name="Complete" category="Maintain" priority="Medium">\n    <definition>' + gdpDef + '</definition>\n    <fit_criterion>None</fit_criterion>\n    <issue>None</issue>\n  </goal_environment>\n</goal>\n'

    for obsName in obsNames.split('#'):
      gaBuf += '<goal_association environment="Complete" goal_name=\"' + obsName + '\" goal_dim="obstacle" ref_type="resolve" subgoal_name=\"' + gdpName + '\" subgoal_dim=\"'
      if isSatisfied == 'Y':
        gaBuf += 'goal'
      else:
        gaBuf += 'domainproperty'
      gaBuf += '\" alternative_id="0">\n  <rationale>' + rowDict['Rationale'] + '</rationale>\n</goal_association>\n'
      aaBuf += '| ' + obsName + ' | ' + gdpName + ' | ' + gdpDef + ' | ' + cNames + ' | ' + isSatisfied + ' | ' + dpDef + ' |\n'

  if len(dpBuf) > 0:
    xmlBuf += '<goals>\n' + dpBuf + gBuf + '</goals>\n\n'
  if len(gaBuf) > 0:
    xmlBuf += '<associations>\n' + gaBuf + '</associations>\n\n'
  xmlBuf += '</cairis_model>'

  outputFile = outputDir + '/pattern_mitigation.xml'
  f = open(outputFile,'w')
  f.write(xmlBuf)
  f.close() 

  aaFile = outputDir + '/ambiguityAnalysis.txt'
  f = open(aaFile,'w')
  f.write(aaBuf)
  f.close() 

  print 'Exported pattern mitigation details to XML'
